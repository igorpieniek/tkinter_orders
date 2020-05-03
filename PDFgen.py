from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Border, Side, Alignment, Font
from tkinter import messagebox
import datetime

class PDFgen():
    def __init__(self):
         self.ex = Workbook()  
         self.er = self.ex.active

         self.__startArrayLine = 4
         self.er.column_dimensions['A'].width = 20
         self.er.row_dimensions[1].height=30
         self.er.column_dimensions['D'].width = 12
         self.er.column_dimensions['E'].width = 34

         self._restFont = Font(size=14 )
         self._companyFont = Font(size = 18, bold =True)

    def process(self, order, mode = 'NORMAL'):
        if self.__readPath():
            self.__setMode(mode)
            self._clearExcelSheet()
            self.__copyProducts(order)   
            self._addDefaultCells()
            #self._delEmptyLines()
            self._addToExcel()
            self._generatePDF()

    def __setMode(self, mode):
        self.__mode = mode
        if self.__mode == 'PRODUCTION' or self.__mode == 'SLEEVES':
            self.__startArrayLine = 3

    def _clearExcelSheet(self):
        for row in self.er['A1':'G37']:
            for cell in row:
                cell.value = None

    def __copyProducts(self, order):
        self.__orderOBJ = order
        self.__order = order.getOrderDict()
        
    def _addDefaultCells(self):
        self.er['A' + str(self.__startArrayLine)] = 'Model'
        self.er['B' + str(self.__startArrayLine)] = 'Suma'
        self.er['C' + str(self.__startArrayLine)] = 'Ilość'
        self.er['D' + str(self.__startArrayLine)] = 'Rodzaj'
        self.er['E' + str(self.__startArrayLine)] = 'Uwagi'

        headerFont = Font(size=16, bold= True )
        for row in self.er['A'+ str(self.__startArrayLine):'E'+ str(self.__startArrayLine)]:
            for cell in row:
                cell.font = headerFont

        bd = Side(style='thin', color="000000")
        self.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        self.__addCompanyName() # add company name to
        self.__addOrderDate()
        if  self.__mode == 'NORMAL':
            self.__addCollectDate()
            self.__addPayment()
            self.__addInvoice()


    def __addCompanyName(self):
        self.er.merge_cells('A1:B1')
        self.er['A1'] =  self.__order['companyName']
        self.er['A1'].alignment = Alignment(horizontal='left', vertical='center')
        cell = self.er['A1']
        cell.font = self._companyFont

    def __addOrderDate(self):
        self.__addDates(infoCell = 'A2', valueCell = 'B2',mergeCell = 'C2', 
                           infoMsg = 'Data zamówienia:', date = self.__order['dateOrder'] )

    def __addCollectDate(self):
        self.__addDates(infoCell = 'A3', valueCell = 'B3',mergeCell = 'C3', 
                           infoMsg = 'Data odbioru:', date = self.__order['dateCollect'] )


    def __addDates(self,*,infoCell, valueCell,mergeCell, infoMsg, date,fontSize=12):
        self.er[infoCell] =  infoMsg
        self.er[infoCell].alignment = Alignment(horizontal='left', vertical='center')
        cell = self.er[infoCell]
        cell.font =  Font(size=fontSize )

        self.er.merge_cells(valueCell + ':' + mergeCell)
        if date == datetime.date.fromordinal(1) : self.er[ valueCell] = '-------'
        else: self.er[ valueCell ] = str(date.day)+'.'+str(date.month)+'.'+str(date.year)
        self.er[ valueCell ].alignment = Alignment(horizontal='left', vertical='center')
        cell = self.er[  valueCell ]
        cell.font =  Font(size=fontSize )

    def __addPayment(self):
        self.__addInvoiceAndPayment(infoCell = 'D3', valueCell = 'E3', 
                                     infoMsg = 'Płatność: ',valueMsg = self.__order['payment'] )

    def __addInvoice(self):
        self.__addInvoiceAndPayment(infoCell = 'D2', valueCell = 'E2', 
                                     infoMsg = 'Nr faktury: ',valueMsg = self.__order['invoice'] )


    def __addInvoiceAndPayment(self,*,infoCell, valueCell, infoMsg, valueMsg, fontSize=12):
        self.er[infoCell] =  infoMsg
        self.er[infoCell].alignment = Alignment(horizontal='left', vertical='center')
        cell = self.er[infoCell]
        cell.font =  Font(size= fontSize )

        if  valueMsg == '' or valueMsg==0: self.er[valueCell] = '-------'
        else: self.er[valueCell] =  valueMsg
        self.er[valueCell].alignment = Alignment(horizontal='left', vertical='center')
        cell = self.er[valueCell]
        cell.font =  Font(size= fontSize )

    def _addToExcel(self):
        lastRow = self.__startArrayLine + 1

        for model,products in self.__order['products'].items():
                if not products: continue
                self.er['A'+ str(lastRow)] = products[0].getModel()
                self.er['B'+ str(lastRow)] = self.__order['sum'][model]
                self.er['A'+ str(lastRow)].alignment = Alignment(horizontal='center', vertical='center',wrap_text = True)
                self.er['B'+ str(lastRow)].alignment = Alignment(horizontal='center', vertical='center')
                k =0
                while k < len(products):
                    self.er['D'+ str(k+lastRow)] = products[k].getKind()
                    self.er['C'+ str(k+lastRow)] = products[k].getNumber()
                    self.er['C'+ str(k+lastRow)].alignment = Alignment(horizontal='right', vertical='center')
                    self.er['D'+ str(k+lastRow)].alignment = Alignment(horizontal='center', vertical='center')
                    k+=1
                self.er.merge_cells('B'+ str(lastRow) +':'+ 'B'+str(lastRow -1+ len(products )))
                self.er.merge_cells('A'+ str(lastRow) +':'+ 'A'+str(lastRow -1+ len(products )))
                lastRow += len(products)
        

      
        for row in self.er['A'+ str(self.__startArrayLine):'E'+str(lastRow-1)]:
            for cell in row:
                cell.border = self.border
        for row in self.er['A'+ str(self.__startArrayLine + 1) :'E'+str(lastRow-1)]:
            for cell in row:
                cell.font = self._restFont    
        
        self.ex.save('tkinter_test.xlsx')


    def _generatePDF(self):
        import win32com.client
        from pathlib import Path

        pdfName = str(self.__order['dateOrder'].day) + '-' + str(self.__order['dateOrder'].month) + '-' + str(self.__order['dateOrder'].year)+ self.__order['companyName']

        o = win32com.client.Dispatch("excel.application")
        o.Visible = False
        wb_path = r'c:\users\igor\source\repos\tkinterproject1\tkinterproject1\tkinter_test.xlsx'
        wb = o.Workbooks.Open(wb_path)
        ws_index_list = [1] 
        tempFolderPath =  self.__folderPath + str(self.__order['dateOrder'].year)+ '-'+ str(self.__order['dateOrder'].month) 
        Path( tempFolderPath ).mkdir(parents=True, exist_ok=True)
        path_to_pdf = tempFolderPath + '\\'+ pdfName +'.pdf'
        path_to_pdf = r''.join(path_to_pdf)
        wb.WorkSheets(ws_index_list).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
        wb.Close(True, wb_path)

    def __readPath(self):
        try: 
            file = open('config.txt')
            self.__folderPath = file.read()
            file.close()
        except: 
            print('So such file!')
            messagebox.showerror('Błąd!', 'Nie można wygenerować pliku PDF!\n Najpierw wybierz folder korzystając z ustawień')
            return False
        else: return True